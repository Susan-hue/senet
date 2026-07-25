"""Read-only document exports for a result sheet.

Both builders read the *current* (``is_current``) score rows and never mutate
anything. They are pure functions of a ``CourseResult`` so they can be called
from the authorized-role views, the Celery worker, and the NUC auditor path
alike, with the caller responsible for scoping/authorization.
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.pdfencrypt import StandardEncryption
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from results.models import ResultStatus, StudentScore

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_CONTENT_TYPE = "application/pdf"


def _fmt(value):
    return f"{Decimal(value):.2f}"


def current_scores(result):
    """The current score rows for a result, ordered for a stable report."""
    return list(
        StudentScore.all_objects.filter(result=result, is_current=True)
        .select_related("student")
        .order_by("student__identifier", "student__full_name")
    )


def ogr_context(result):
    """The structured data an OGR is rendered from. Exposed on its own so the
    exact figures can be asserted in tests without parsing a binary PDF."""
    rows = current_scores(result)
    return {
        "institution": result.institution.name,
        "course_code": result.course.code,
        "course_title": result.course.title,
        "session": result.session.name,
        "semester": result.semester.name,
        "lecturer": result.lecturer.full_name,
        "status_label": result.get_status_display(),
        "status_value": result.status,
        "is_ratified": result.status == ResultStatus.RATIFIED_BY_SENATE,
        "students": [
            {
                "matric": row.student.identifier or "—",
                "name": row.student.full_name,
                "ca": _fmt(row.ca_score),
                "exam": _fmt(row.exam_score),
                "total": _fmt(row.total),
                "grade": row.grade,
            }
            for row in rows
        ],
    }


def export_basename(result, extension):
    code = result.course.code.replace(" ", "")
    session = result.session.name.replace("/", "-")
    return f"{code}_{session}_{result.semester.name}.{extension}".replace(" ", "")


def build_ogr_pdf(result):
    """Render a course result as a flat, un-editable Official Grade Report PDF.

    The document is owner-password encrypted with modification disabled, so
    compliant readers treat it as read-only; there are no form fields, so the
    content itself is static."""
    context = ogr_context(result)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)
    name_cell = ParagraphStyle("name", parent=cell, alignment=0)

    buffer = io.BytesIO()
    # Modification and annotation are disabled; the file opens without a password
    # but cannot be edited in a compliant viewer. A random owner password means
    # nobody holds the key to lift those restrictions.
    encryption = StandardEncryption(
        userPassword="",
        ownerPassword=None,
        canModify=0,
        canCopy=1,
        canPrint=1,
        canAnnotate=0,
    )
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=f"Official Grade Report — {context['course_code']}",
        author=context["institution"],
        subject="Official Grade Report",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        encrypt=encryption,
    )

    story = [
        Paragraph(context["institution"], styles["Title"]),
        Paragraph("Official Grade Report", styles["Heading2"]),
        Spacer(1, 6),
        Paragraph(f"<b>{context['course_code']}</b> — {context['course_title']}", styles["Normal"]),
        Paragraph(
            f"Session: {context['session']} &nbsp;&nbsp; Semester: {context['semester']}",
            styles["Normal"],
        ),
        Paragraph(f"Status: <b>{context['status_label']}</b>", styles["Normal"]),
    ]
    if not context["is_ratified"]:
        story.append(
            Paragraph(
                "<b>NOT YET RATIFIED BY SENATE — PROVISIONAL, NOT FOR OFFICIAL USE</b>",
                ParagraphStyle("warn", parent=styles["Normal"], textColor=colors.red),
            )
        )
    story.append(Spacer(1, 10))

    header = ["S/N", "Matric No.", "Name", "CA", "Exam", "Total", "Grade"]
    data = [header]
    for index, student in enumerate(context["students"], start=1):
        data.append(
            [
                str(index),
                Paragraph(student["matric"], cell),
                Paragraph(student["name"], name_cell),
                student["ca"],
                student["exam"],
                student["total"],
                student["grade"],
            ]
        )
    if not context["students"]:
        data.append(["", Paragraph("No scores recorded.", cell), "", "", "", "", ""])

    table = Table(
        data,
        colWidths=[10 * mm, 30 * mm, 62 * mm, 16 * mm, 16 * mm, 16 * mm, 16 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9ca3af")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 26))
    story.append(_signature_block())

    doc.build(story)
    return buffer.getvalue()


def _signature_block():
    line = ParagraphStyle("sig", parent=getSampleStyleSheet()["Normal"], fontSize=9, leading=12)
    cells = [
        [Paragraph("_______________________", line)] * 3,
        [
            Paragraph("<b>Lecturer</b><br/>Name &amp; Signature / Date", line),
            Paragraph("<b>Head of Department</b><br/>Name &amp; Signature / Date", line),
            Paragraph("<b>Dean</b><br/>Name &amp; Signature / Date", line),
        ],
    ]
    table = Table(cells, colWidths=[58 * mm, 58 * mm, 58 * mm])
    table.setStyle(
        TableStyle(
            [
                ("TOPPADDING", (0, 0), (-1, 0), 2),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 2),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


# --------------------------------------------------------------------------- #
# Broadsheet (.xlsx)                                                           #
# --------------------------------------------------------------------------- #

_HEADER = ["S/N", "Matric No.", "Name", "CA", "Exam", "Total", "Grade"]


def build_broadsheet_workbook(result):
    """A single-sheet broadsheet: the whole class as rows with CA/Exam/Total/
    Grade columns, laid out for Senate consideration."""
    context = ogr_context(result)
    wb = Workbook()
    ws = wb.active
    ws.title = "Broadsheet"

    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = len(_HEADER)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = context["institution"]
    ws["A1"].font = title_font
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{context['course_code']} — {context['course_title']}"
    ws["A2"].font = meta_font
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = (
        f"Session: {context['session']}    Semester: {context['semester']}    "
        f"Status: {context['status_label']}"
    )
    ws["A3"].font = meta_font

    header_row = 5
    for col, label in enumerate(_HEADER, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for offset, student in enumerate(context["students"], start=1):
        row = header_row + offset
        values = [
            offset,
            student["matric"],
            student["name"],
            float(student["ca"]),
            float(student["exam"]),
            float(student["total"]),
            student["grade"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col != 3:
                cell.alignment = center

    widths = [6, 18, 34, 8, 8, 8, 8]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return wb


def build_broadsheet_xlsx(result):
    buffer = io.BytesIO()
    build_broadsheet_workbook(result).save(buffer)
    return buffer.getvalue()
