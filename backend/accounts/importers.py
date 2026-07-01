"""Bulk import for students and courses (.csv and .xlsx).

Pattern: validate every row first; invalid or duplicate rows are skipped and
reported row-by-row; the valid subset is inserted in a single atomic
transaction (all-or-nothing for that subset). One bad row never blocks the rest.

The importer takes an explicit ``institution`` so it behaves identically whether
called synchronously from a view or asynchronously from a Celery worker (where
the request-scoped tenant thread-local is not set). Uploads are normalised to
CSV text up front (see ``decode_upload``) so the engine and the Celery task
stay text-based regardless of the original file format.
"""

import csv
import io
from dataclasses import dataclass, field

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models.functions import Upper
from openpyxl import load_workbook

from accounts.models import (
    Course,
    CourseAssignment,
    Department,
    Level,
    Role,
    Semester,
    Session,
    User,
)

STUDENT_REQUIRED_COLUMNS = [
    "full_name",
    "email",
    "matric_number",
    "department_code",
    "current_level",
]
COURSE_REQUIRED_COLUMNS = ["code", "title", "credit_units", "level", "department_code"]
# lecturer_identifier is an optional fallback used when lecturer_email is blank.
ASSIGNMENT_REQUIRED_COLUMNS = ["lecturer_email", "course_code", "session", "semester"]

_LEVEL_VALUES = sorted(Level.values)


class ImportFileError(Exception):
    """A file-level problem (empty, undecodable, missing required columns)."""


def _cell_to_text(cell):
    """Render an openpyxl cell as text, keeping whole numbers integer-clean.

    Excel stores numbers as floats, so a credit_units of ``3`` arrives as
    ``3.0``; emit ``"3"`` so downstream integer parsing succeeds.
    """
    if cell is None:
        return ""
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def _xlsx_to_csv_text(raw_bytes):
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure means an unreadable file
        raise ImportFileError("The .xlsx file could not be read.") from exc

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for values in workbook.active.iter_rows(values_only=True):
        writer.writerow([_cell_to_text(cell) for cell in values])
    workbook.close()
    return buffer.getvalue()


def decode_upload(filename, raw_bytes):
    """Normalise an uploaded .csv or .xlsx (as bytes) into CSV text.

    Raises ImportFileError for undecodable CSV or unreadable spreadsheets.
    """
    if (filename or "").lower().endswith(".xlsx"):
        return _xlsx_to_csv_text(raw_bytes)
    try:
        return raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileError("File must be a UTF-8 encoded CSV.") from exc


@dataclass
class ImportResult:
    total: int = 0
    created: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)

    @property
    def summary(self):
        return {
            "total_rows": self.total,
            "created": self.created,
            "skipped": self.skipped,
        }

    @property
    def message(self):
        return f"{self.created} of {self.total} row(s) imported, {self.skipped} skipped."


def _read_rows(text, required_columns):
    """Parse CSV text into ``[(row_number, {column: value}), ...]``.

    Row numbers are 1-based spreadsheet rows (header is row 1, first data row 2).
    Raises ImportFileError for file-level problems.
    """
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ImportFileError("The file is empty.") from exc

    header = [(column or "").strip().lower() for column in header]
    missing = [column for column in required_columns if column not in header]
    if missing:
        raise ImportFileError(f"Missing required column(s): {', '.join(missing)}.")

    rows = []
    for number, values in enumerate(reader, start=2):
        if not any((value or "").strip() for value in values):
            continue  # ignore fully blank lines
        data = {
            header[index]: (values[index].strip() if index < len(values) else "")
            for index in range(len(header))
        }
        rows.append((number, data))
    return rows


def _parse_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _departments_by_code(institution):
    """Map upper-cased department code -> department for case-insensitive lookup.

    The stored code keeps its canonical casing; only the lookup key is folded.
    """
    return {
        department.code.upper(): department
        for department in Department.all_objects.filter(institution=institution)
    }


# --------------------------------------------------------------------------- #
# Students                                                                    #
# --------------------------------------------------------------------------- #


def import_students(institution, text):
    rows = _read_rows(text, STUDENT_REQUIRED_COLUMNS)
    departments = _departments_by_code(institution)

    emails = [r["email"].strip().lower() for _, r in rows if r.get("email", "").strip()]
    matrics = [r["matric_number"].strip() for _, r in rows if r.get("matric_number", "").strip()]
    existing_emails = set(User.objects.filter(email__in=emails).values_list("email", flat=True))
    existing_matrics = set(
        User.objects.filter(institution=institution, identifier__in=matrics).values_list(
            "identifier", flat=True
        )
    )

    seen_emails, seen_matrics = set(), set()
    to_create, errors = [], []

    for number, data in rows:
        row_errors = []
        full_name = data.get("full_name", "").strip()
        email = data.get("email", "").strip().lower()
        matric = data.get("matric_number", "").strip()
        department_code = data.get("department_code", "").strip()
        level_raw = data.get("current_level", "").strip()

        if not full_name:
            row_errors.append("full_name is required")

        if not email:
            row_errors.append("email is required")
        else:
            try:
                validate_email(email)
            except DjangoValidationError:
                row_errors.append(f"email '{email}' is not valid")

        if not matric:
            row_errors.append("matric_number is required")

        department = departments.get(department_code.upper()) if department_code else None
        if not department_code:
            row_errors.append("department_code is required")
        elif department is None:
            row_errors.append(f"department_code '{department_code}' not found")

        level = _parse_int(level_raw)
        if not level_raw:
            row_errors.append("current_level is required")
        elif level is None:
            row_errors.append(f"current_level '{level_raw}' must be a number")
        elif level not in Level.values:
            row_errors.append(f"current_level '{level_raw}' must be one of {_LEVEL_VALUES}")

        if email and email in seen_emails:
            row_errors.append("duplicate email in file")
        elif email and email in existing_emails:
            row_errors.append(f"email '{email}' already exists")

        if matric and matric in seen_matrics:
            row_errors.append("duplicate matric_number in file")
        elif matric and matric in existing_matrics:
            row_errors.append(f"duplicate matric_number '{matric}'")

        if row_errors:
            errors.append({"row": number, "errors": row_errors})
            continue

        seen_emails.add(email)
        seen_matrics.add(matric)
        student = User(
            email=email,
            full_name=full_name,
            role=Role.STUDENT,
            institution=institution,
            department=department,
            current_level=level,
            identifier=matric,
            is_verified=False,
            is_active=True,
        )
        student.set_unusable_password()
        to_create.append(student)

    with transaction.atomic():
        User.objects.bulk_create(to_create, batch_size=500)

    return ImportResult(total=len(rows), created=len(to_create), skipped=len(errors), errors=errors)


# --------------------------------------------------------------------------- #
# Courses                                                                      #
# --------------------------------------------------------------------------- #


def import_courses(institution, text):
    rows = _read_rows(text, COURSE_REQUIRED_COLUMNS)
    departments = _departments_by_code(institution)

    codes_upper = [r["code"].strip().upper() for _, r in rows if r.get("code", "").strip()]
    existing_codes = set(
        Course.all_objects.filter(institution=institution)
        .annotate(code_upper=Upper("code"))
        .filter(code_upper__in=codes_upper)
        .values_list("code_upper", flat=True)
    )

    seen_codes = set()
    to_create, errors = [], []

    for number, data in rows:
        row_errors = []
        code = data.get("code", "").strip()
        title = data.get("title", "").strip()
        units_raw = data.get("credit_units", "").strip()
        level_raw = data.get("level", "").strip()
        department_code = data.get("department_code", "").strip()
        ca_raw = data.get("ca_weight", "").strip()
        exam_raw = data.get("exam_weight", "").strip()

        if not code:
            row_errors.append("code is required")
        if not title:
            row_errors.append("title is required")

        units = _parse_int(units_raw)
        if not units_raw:
            row_errors.append("credit_units is required")
        elif units is None:
            row_errors.append(f"credit_units '{units_raw}' must be a number")
        elif units <= 0:
            row_errors.append("credit_units must be greater than zero")

        level = _parse_int(level_raw)
        if not level_raw:
            row_errors.append("level is required")
        elif level is None or level not in Level.values:
            row_errors.append(f"level '{level_raw}' must be one of {_LEVEL_VALUES}")

        department = departments.get(department_code.upper()) if department_code else None
        if not department_code:
            row_errors.append("department_code is required")
        elif department is None:
            row_errors.append(f"department_code '{department_code}' not found")

        ca = exam = None
        has_ca, has_exam = bool(ca_raw), bool(exam_raw)
        if has_ca or has_exam:
            if not (has_ca and has_exam):
                row_errors.append("provide both ca_weight and exam_weight, or neither")
            else:
                ca, exam = _parse_int(ca_raw), _parse_int(exam_raw)
                if ca is None or exam is None:
                    row_errors.append("ca_weight and exam_weight must be numbers")
                elif ca + exam != 100:
                    row_errors.append("ca_weight and exam_weight must sum to 100")

        code_key = code.upper()
        if code and code_key in seen_codes:
            row_errors.append("duplicate code in file")
        elif code and code_key in existing_codes:
            row_errors.append(f"course code '{code}' already exists")

        if row_errors:
            errors.append({"row": number, "errors": row_errors})
            continue

        seen_codes.add(code_key)
        weights_set = has_ca and has_exam
        to_create.append(
            Course(
                institution=institution,
                department=department,
                code=code,
                title=title,
                credit_units=units,
                level=level,
                ca_weight=ca if weights_set else None,
                exam_weight=exam if weights_set else None,
            )
        )

    with transaction.atomic():
        Course.all_objects.bulk_create(to_create, batch_size=500)

    return ImportResult(total=len(rows), created=len(to_create), skipped=len(errors), errors=errors)


# --------------------------------------------------------------------------- #
# Course assignments (lecturer -> course, per term)                           #
# --------------------------------------------------------------------------- #


def import_assignments(institution, text):
    """Bulk-assign lecturers to courses for a session + semester.

    Admin-only in the API; the HOD department scope of the single-assignment
    endpoint does not apply here.
    """
    rows = _read_rows(text, ASSIGNMENT_REQUIRED_COLUMNS)

    lecturers = User.objects.filter(institution=institution, role=Role.LECTURER)
    by_email = {lecturer.email.lower(): lecturer for lecturer in lecturers}
    by_identifier = {u.identifier: u for u in lecturers if u.identifier}
    courses = {c.code.upper(): c for c in Course.all_objects.filter(institution=institution)}
    sessions = {s.name.upper(): s for s in Session.all_objects.filter(institution=institution)}
    semesters = {
        (sem.session_id, sem.name.upper()): sem
        for sem in Semester.all_objects.filter(institution=institution)
    }

    existing = set(
        CourseAssignment.all_objects.filter(institution=institution).values_list(
            "lecturer_id", "course_id", "session_id", "semester_id"
        )
    )
    seen = set()
    to_create, errors = [], []

    for number, data in rows:
        row_errors = []
        email = data.get("lecturer_email", "").strip().lower()
        identifier = data.get("lecturer_identifier", "").strip()
        course_code = data.get("course_code", "").strip()
        session_name = data.get("session", "").strip()
        semester_name = data.get("semester", "").strip()

        lecturer = by_email.get(email) if email else by_identifier.get(identifier)
        if not email and not identifier:
            row_errors.append("lecturer_email or lecturer_identifier is required")
        elif lecturer is None:
            row_errors.append(f"lecturer '{email or identifier}' not found or is not a lecturer")

        course = courses.get(course_code.upper()) if course_code else None
        if not course_code:
            row_errors.append("course_code is required")
        elif course is None:
            row_errors.append(f"course_code '{course_code}' not found")

        session = sessions.get(session_name.upper()) if session_name else None
        if not session_name:
            row_errors.append("session is required")
        elif session is None:
            row_errors.append(f"session '{session_name}' not found")

        semester = None
        if not semester_name:
            row_errors.append("semester is required")
        elif session is not None:
            semester = semesters.get((session.id, semester_name.upper()))
            if semester is None:
                row_errors.append(
                    f"semester '{semester_name}' not found in session '{session_name}'"
                )

        if lecturer and course and session and semester:
            key = (lecturer.id, course.id, session.id, semester.id)
            if key in seen:
                row_errors.append("duplicate assignment in file")
            elif key in existing:
                row_errors.append("assignment already exists")

            if not row_errors:
                seen.add(key)
                to_create.append(
                    CourseAssignment(
                        institution=institution,
                        lecturer=lecturer,
                        course=course,
                        session=session,
                        semester=semester,
                    )
                )

        if row_errors:
            errors.append({"row": number, "errors": row_errors})

    with transaction.atomic():
        CourseAssignment.all_objects.bulk_create(to_create, batch_size=500)

    return ImportResult(total=len(rows), created=len(to_create), skipped=len(errors), errors=errors)
